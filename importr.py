"""
#
#   TEMPLATE
#   The Template-class writes the info to the proper excel-template.
#
#
#   Attributes:
#       DATE_FORMAT (str): Date format used for import
"""
DATE_FORMAT = "%d.%m.%Y"

import debug

import openpyxl

import os
from os.path import expanduser
from datetime import datetime

from ui import helper
from core.obj import proj, arch, corp

from PyQt5.QtCore import QDate


class KFImporter:
    """Import *-Kostenfortschreibung.xlsx files/projects.

    Attributes:
        companies (list): List of Company objects
        companies_args (list): List of dicts
        file_path (pathlib.Path): Path to xlsx-file
        identifier (str): Identifier of the project
        invoices (list): List of Invoice objects
        invoices_args (list):  List of dicts
        jobs (list): List of Job objects
        jobs_args (list):  List of dicts
        trades (list): List of Trade objects
        trades_args (list):  List of dicts
        wb (openpyxl.Workbook): Excel workbook
    """

    def __init__(self, file_path):
        """Initilize KFImporter.

        Args:
            file_path (str): Path to xlsx-file
        """
        self.file_path = pathlib.Path(file_path)

        self.identifier = file_path.name.split("-")[0]

        self.wb = openpyxl.load_workbook(self.file_path, data_only=True)

        self.invoices_args = list()
        self.jobs_args = list()
        self.trades_args = list()
        self.companies_args = list()

        self.invoices = list()
        self.jobs = list()
        self.trades = list()
        self.companies = list()

    def import_data(self):
        """Do the import."""
        self.import_trades_args()
        self.import_companies_args()
        self.import_jobs_args()
        self.import_invoices_args()

    def import_trades_args(self):
        """Improt the trades from the worksheet."""
        ws = self.wb["Gewerke"]
        for i in range(2, len(ws["A"])):
            args = {
                "name": ws[f"A{i}"].value,
                "budget": ws[f"C{i}"].value if ws[f"C{i}"].value else 0,
                "comment": ws[f"E{i}"].value if ws[f"E{i}"].value else "",
            }
            if args["name"]:
                self.trades_args.append(args)

    def import_companies_args(self):
        """Improt the companies from the worksheet."""
        ws = self.wb["Firmen"]
        for i in range(2, len(ws["A"])):
            args = {
                "name": ws[f"A{i}"].value,
                "service": ws[f"B{i}"].value,
                "service_type": ws[f"C{i}"].value,
            }
            if args["name"]:
                self.companies_args.append(args)

    def import_jobs_args(self):
        """Improt the jobs from the worksheet."""
        ws = self.wb["Aufträge"]
        for i in range(2, len(ws["A"])):
            args = {
                "id": int(ws[f"A{i}"].value.split(" ")[1]) if ws[f"A{i}"].value else 0,
                "company_ref": {"name": ws[f"B{i}"].value},
                "cost_group_ref": {
                    "id": str(ws[f"D{i}"].value),
                },
                "trade_ref": {"name": ws[f"C{i}"].value},
                "job_sum": float(ws[f"E{i}"].value) if ws[f"E{i}"].value else 0,
                "comment": ws[f"G{i}"].value if ws[f"G{i}"].value else "",
            }
            if args["id"] != 0:
                self.jobs_args.append(args)

    def import_invoices_args(self):
        """Improt the invoices from the worksheet."""
        ws = self.wb["Kostenfortschreibung"]
        for i in range(7, len(ws["A"])):
            invoice_date = (
                ws[f"H{i}"].value if ws[f"H{i}"].value else str2date("01.07.2020")
            )
            start_vat_reduction_date = str2date("01.07.2020")
            end_vat_reduction_date = str2date("31.12.2020")
            VAT = ws[f"V{i}"].value
            if VAT == 1:
                if start_vat_reduction_date <= invoice_date <= end_vat_reduction_date:
                    VAT = 0.16
                else:
                    VAT = 0.19

            args = {
                "id": ws[f"A{i}"].value,
                "cumulative": ws[f"C{i}"].value,
                "company_ref": {
                    "name": ws[f"D{i}"].value,
                },
                "job_ref": {
                    "id": int(ws[f"G{i}"].value.split(" ")[1])
                    if ws[f"G{i}"].value
                    else "",
                    "company.name": ws[f"D{i}"].value if ws[f"D{i}"].value else "",
                },
                "invoice_date": QDate(
                    invoice_date.year, invoice_date.month, invoice_date.day
                )
                if invoice_date
                else None,
                "inbox_date": QDate(
                    ws[f"I{i}"].value.year,
                    ws[f"I{i}"].value.month,
                    ws[f"I{i}"].value.day,
                )
                if ws[f"I{i}"].value
                else None,
                "checked_date": QDate(
                    ws[f"J{i}"].value.year,
                    ws[f"J{i}"].value.month,
                    ws[f"J{i}"].value.day,
                )
                if ws[f"J{i}"].value
                else None,
                "amount": ws[f"K{i}"].value if ws[f"K{i}"].value else 0,
                "verified_amount": ws[f"L{i}"].value if ws[f"L{i}"].value else 0,
                "rebate": ws[f"N{i}"].value if ws[f"N{i}"].value else 0,
                "reduction_insurance_costs": ws[f"P{i}"].value
                if ws[f"P{i}"].value
                else 0,
                "reduction_usage_costs": ws[f"R{i}"].value if ws[f"R{i}"].value else 0,
                "VAT": VAT,
                "safety_deposit": ws[f"Y{i}"].value if ws[f"Y{i}"].value else 0,
                "discount": ws[f"AB{i}"].value if ws[f"AB{i}"].value else 0,
                "due_date": QDate(1, 1, 1),
                "due_date_discount": QDate(1, 1, 1) if ws[f"AB{i}"].value else None,
            }
            if args["id"]:
                self.invoices_args.append(args)

    def create_objects(self):
        """Create the ojects from the imported args."""
        for trade_args in self.trades_args:
            trade = arch.Trade(**trade_args)
            self.trades.append(trade)

        for company_args in self.companies_args:
            company = corp.Company(**company_args)
            self.companies.append(company)

        for job_args in self.jobs_args:
            job = arch.ArchJob(**job_args)
            self.jobs.append(job)

        for invoice_args in self.invoices_args:
            invoice = corp.Invoice(**invoice_args)
            self.invoices.append(invoice)


def str2date(date_str):
    """Convert String to Date.

    Args:
        date_str (str): Input

    Returns:
        datetime.datetime.Date: Output
    """
    return datetime.strptime(date_str, DATE_FORMAT)


def date2str(date):
    """Convert Date to String.

    Args:
        date (datetime.datetime.Date): Input

    Returns:
        str: Output
    """
    return date.strftime(DATE_FORMAT)
