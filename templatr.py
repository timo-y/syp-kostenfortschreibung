"""
#
#   TEMPLATE
#   The Template-class writes the info to the proper excel-template.
#   The templates are stored in the template folder. Every class writes
#   the data in the corresponding cell of the worksheet.
#
"""
import debug

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side

import os
from os.path import expanduser

from ui import helper


class Template:
    """Template wrapper class with basic functionallity.

    Attributes:
        align_center (openpyxl.styles.Alignment): Center cell-alignment
        align_left (openpyxl.styles.Alignment):  Left cell-alignment
        align_right (openpyxl.styles.Alignment):  Right cell-alignment
        amount_format (str): Format of the amount-cells
        bold_font_header (openpyxl.styles): Header font
        bold_font_table_header (openpyxl.styles.Border): Table header font
        date_format (str): Format of the date-cells
        normal_font (openpyxl.styles.Font): Standard font style
        number_format (str): Format of the number-cells
        percent_format (str): Format of the percent-cells
        template_dir (Path): Path to template-directory
        template_file_path (Path): Path to template-file
        wb (TYPE): The workbook of the template
        ws (TYPE): The worksheet of the workbook
    """

    def __init__(self, template_dir, template_filename, save_path):
        """Initialize Template.

        Args:
            template_dir (Path): Path to template directory
            template_filename (str): Template filename
            save_path (Path): Path for output
        """
        self.template_dir = template_dir
        self.template_file_path = os.path.join(self.template_dir, template_filename)

        self.wb = openpyxl.load_workbook(self.template_file_path)
        self.ws = self.wb["TEMPLATE"]

        self.date_format = "DD.MM.YYYY"
        self.number_format = "#,##0.00 ;-#,##0.00 ; - "
        self.amount_format = "#,##0.00 € ;-#,##0.00 € ; - € "
        self.percent_format = "0.0% ;0.0% ; - % "

        self.normal_font = Font(color="000000", bold=False, size=12)
        self.bold_font_header = Font(color="000000", bold=True, size=18)
        self.bold_font_table_header = Font(color="000000", bold=True, size=12)

        self.align_right = Alignment(horizontal="right")
        self.align_center = Alignment(horizontal="center")
        self.align_left = Alignment(horizontal="left")

    @debug.log
    def add_rows(self, before_row, number_of_rows):
        """Add rows in a template before some row.
        Add rows, by simply moving "everything" (A-AA and 99 rows) down by
        the amount of rows you want to add

        Args:
            before_row (int): Row index where to insert the new rows
            number_of_rows (int): Number of rows to insert
        """
        self.ws.move_range(
            f"A{before_row}:AA{99+before_row}", rows=number_of_rows, cols=0
        )

    @debug.log
    def make_cells(self, excel_data):
        """Write data in the cells.

        Args:
            excel_data (dict): Data to write
        """
        for write_cell in excel_data:
            # Data can be assigned directly to cells
            if "data" in write_cell.keys():
                self.ws[write_cell["cell"]] = write_cell["data"]
            if "font" in write_cell.keys():
                self.ws[write_cell["cell"]].font = write_cell["font"]
            if "alignment" in write_cell.keys():
                self.ws[write_cell["cell"]].alignment = write_cell["alignment"]
            if "number_format" in write_cell.keys():
                self.ws[write_cell["cell"]].number_format = write_cell["number_format"]
            if "border" in write_cell.keys():
                self.ws[write_cell["cell"]].border = write_cell["border"]
            if "row_height" in write_cell.keys():
                cell = self.ws[write_cell["cell"]]
                self.ws.row_dimensions[cell.row].height = write_cell["row_height"]

    @debug.log
    def save_file(self):
        """Save to a file at save_path."""
        # create directory if non-existing
        if not os.path.exists(os.path.dirname(self.save_path)):
            os.makedirs(os.path.dirname(self.save_path))
        # Save the file
        self.wb.save(self.save_path)


class InvoiceCheckExcelTemplate(Template):
    """Template for the invoice check.

    Attributes:
        excel_data (dict): Data to write
        filename (str): Filename for output
        new_rows (int): Number of added rows for adjusting the printing area
        save_dir (Path): Path to save directory
        save_path (Path): Path to save file
        template_filename (str): Name of the template-file
    """

    def __init__(self, app_data, invoice, save_dir, filename=None):
        """Initialize Template.

        Args:
            app_data (api.AppData): Application data containing the project data
            invoice (corp.Invoice): Invoice for the invoice check
            save_dir (Path): Path to save directory
            filename (None, optional): Filename for output
        """
        self.template_filename = "invoice_check.xlsx"

        self.filename = (
            filename
            if filename
            else f"{helper.today_str()}-invoice_check-{invoice.id.replace(' ', '_')}"
        )
        self.save_dir = save_dir
        self.save_path = os.path.join(
            app_data.get_dir(), self.save_dir, f"{self.filename}.xlsx"
        )

        super(InvoiceCheckExcelTemplate, self).__init__(
            template_dir=app_data.config["template_subdir"],
            template_filename=self.template_filename,
            save_path=self.save_path,
        )
        """
        #   new_rows
        #   If we have more than 10 previous invoices,
        #   we need to add rows and hence have to move
        #   the cells underneath the listing by the number
        #   of lines added
        """
        self.new_rows = 0
        if len(invoice.prev_invoices) > 10:
            self.new_rows = len(invoice.prev_invoices) - 10
            self.add_rows(15, self.new_rows)

        """
        #
        #   EXCEL DATA
        #   Gather the data for the invoice check and connect
        #   it to the cells.
        #
        """
        date = helper.today_str("%d.%m.%Y")
        self.excel_data = [
            {"cell": "D7", "data": invoice.company.name},
            {"cell": "H8", "data": invoice.invoice_date.toPyDate()},
            {"cell": "C8", "data": app_data.project.identifier},
            {"cell": "C10", "data": " ".join(["Auftrag:", str(invoice.job.id)])},
            {"cell": "E11", "data": invoice.id},
            {"cell": "H11", "data": invoice.amount},
            {"cell": "H12", "data": invoice.verified_amount},
            {"cell": f"G{25+self.new_rows}", "data": invoice.reduction_total},
            {"cell": f"H{25+self.new_rows}", "data": invoice.reduction_total_amount},
            {
                "cell": f"H{26+self.new_rows}",
                "data": invoice.amount_a_reductions_amount,
            },
            {"cell": f"C{27+self.new_rows}", "data": invoice.VAT},
            {
                "cell": f"H{27+self.new_rows}",
                "data": invoice.amount_a_reductions_amount_VAT_amount,
            },
            {
                "cell": f"H{28+self.new_rows}",
                "data": invoice.amount_a_reductions_amount_w_VAT,
            },
            {"cell": f"E{31+self.new_rows}", "data": invoice.safety_deposit},
            {"cell": f"H{31+self.new_rows}", "data": invoice.safety_deposit_amount},
            {
                "cell": f"H{32+self.new_rows}",
                "data": invoice.job.paid_safety_deposits_sum,
            },
            {"cell": f"H{34+self.new_rows}", "data": invoice.approved_amount},
            {
                "cell": f"E{35+self.new_rows}",
                "data": invoice.discount if invoice.discount else "-",
            },
            {
                "cell": f"H{35+self.new_rows}",
                "data": invoice.discount_amount if invoice.discount else "-",
            },
            {
                "cell": f"H{36+self.new_rows}",
                "data": invoice.approved_amount_a_discount_amount
                if invoice.discount
                else "-",
            },
            {
                "cell": f"D{39+self.new_rows}",
                "data": invoice.checked_date.toPyDate()
                if invoice.checked_date
                else invoice.invoice_date.toPyDate(),
            },
            {"cell": f"F{34+self.new_rows}", "data": invoice.due_date.toPyDate()},
            {
                "cell": f"F{36+self.new_rows}",
                "data": invoice.due_date_discount.toPyDate()
                if invoice.due_date_discount
                else "-",
            },
        ]
        """
        #
        #   PREVIOUS INVOICES
        #   Add one line per previous invoice.
        #
        """
        j = 0
        for prev_invoice in invoice.prev_invoices:
            invoice_line = [
                {"cell": f"C{15+j}", "border": Border(left=Side(style="thin"))},
                {
                    "cell": f"D{15+j}",
                    "data": prev_invoice.invoice_date.toPyDate(),
                    "number_format": self.date_format,
                },
                {"cell": f"F{15+j}", "data": prev_invoice.id},
                {
                    "cell": f"H{15+j}",
                    "data": prev_invoice.verified_amount,
                    "number_format": self.amount_format,
                    "border": Border(right=Side(style="thin")),
                },
            ]
            self.excel_data.extend(invoice_line)
            j += 1

    @debug.log
    def make_file(self):
        """Make the file."""
        self.make_cells(self.excel_data)
        self.save_file()


class CompanyOVExcelTemplate(Template):
    """Template for an overview of the progress of a company.

    Attributes:
        excel_data (dict): Data to write
        filename (str): Filename for output
        last_row_index (int): Index of the last used row
        save_dir (Path): Path to save directory
        save_path (Path): Path to save file
        template_filename (str): Name of the template-file
    """

    def __init__(self, app_data, company, save_dir, selected_job=None, filename=None):
        """Initialize Template.

        Args:
            app_data (api.AppData): Application data containing the project data
            company (TYPE): Description
            save_dir (Path): Path to save directory
            selected_job (None, optional): If selected, only this job is displayed
            filename (None, optional): Filename for output
        """
        self.template_filename = "company_ov.xlsx"

        job_str = f"-{selected_job.id}" if selected_job is not None else ""
        self.filename = (
            filename
            if filename
            else f"{helper.today_str()}-company_ov-{company.name.replace(' ', '_')}{job_str}"
        )
        self.save_dir = save_dir
        self.save_path = os.path.join(
            app_data.get_dir(), self.save_dir, f"{self.filename}.xlsx"
        )

        self.last_row_index = 9  # last row used to define the printed area

        super(CompanyOVExcelTemplate, self).__init__(
            template_dir=app_data.config["template_subdir"],
            template_filename=self.template_filename,
            save_path=self.save_path,
        )
        """
        #
        #   EXCEL DATA
        #   Gather the data for the invoice check and connect
        #   it to the cells.
        #
        """
        date = helper.today_str("%d.%m.%Y")
        self.excel_data = [
            {"cell": "B3", "data": date, "number_format": self.date_format},
            {"cell": "B4", "data": app_data.project.identifier},
            {
                "cell": "B5",
                "data": app_data.project.client if app_data.project.client else "",
            },
            {"cell": "B7", "data": company.name},
        ]

        """
        #   new_rows
        #   We need one row for each job and each invoice
        #
        """
        titles = app_data.get_titles()
        j = 0
        header_height = 30
        jobs = [selected_job]
        if selected_job is None:
            jobs = app_data.project.get_jobs_of_company(company)
        for job in jobs:
            job_lines = [
                {
                    "cell": f"A{9+j}",
                    "data": f"Auftrag {job.id}",
                    "border": Border(bottom=Side(style="medium")),
                    "font": self.bold_font_header,
                    "row_height": header_height,
                },
                {"cell": f"B{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"C{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"D{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"E{9+j}", "border": Border(bottom=Side(style="medium"))},
                {
                    "cell": f"B{10+j}",
                    "data": titles["trade"],
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"C{10+j}",
                    "data": titles["cost_group"],
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"D{10+j}",
                    "data": titles["job_sum"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"E{10+j}",
                    "data": titles["job_sum_w_VAT"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {"cell": f"B{11+j}", "data": job.trade.name, "font": self.normal_font},
                {
                    "cell": f"C{11+j}",
                    "data": job.cost_group.id,
                    "font": self.normal_font,
                },
                {
                    "cell": f"D{11+j}",
                    "data": job.job_sum_w_additions,
                    "font": self.normal_font,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"E{11+j}",
                    "data": job.job_sum_w_additions * (1 + app_data.project.get_vat()),
                    "font": self.normal_font,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                },
            ]
            self.excel_data.extend(job_lines)
            invoice_header_lines = [
                {
                    "cell": f"A{12+j}",
                    "data": titles["invoices"],
                    "border": Border(bottom=Side(style="thin", color="999999")),
                    "font": self.normal_font,
                    "row_height": header_height,
                },
                {
                    "cell": f"B{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"C{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"D{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"E{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"B{13+j}",
                    "data": titles["id"],
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"C{13+j}",
                    "data": titles["verified_amount_short"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"D{13+j}",
                    "data": titles["safety_deposit_amount"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"E{13+j}",
                    "data": titles["approved_amount"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
            ]
            self.excel_data.extend(invoice_header_lines)
            j += 5
            invoices_of_job = app_data.project.get_invoices_of_job(job)
            verified_amount_sum = 0
            safety_deposit_sum = 0
            approved_amount_sum = 0
            for invoice in invoices_of_job:
                """sums"""
                verified_amount_sum += invoice.verified_amount
                safety_deposit_sum += invoice.safety_deposit_amount
                approved_amount_sum += invoice.approved_amount
                """ write data """
                invoice_line = [
                    {"cell": f"B{9+j}", "data": invoice.id},
                    {
                        "cell": f"C{9+j}",
                        "data": invoice.verified_amount,
                        "number_format": self.amount_format,
                        "alignment": self.align_right,
                    },
                    {
                        "cell": f"D{9+j}",
                        "data": invoice.safety_deposit_amount,
                        "number_format": self.amount_format,
                        "alignment": self.align_right,
                    },
                    {
                        "cell": f"E{9+j}",
                        "data": invoice.approved_amount,
                        "number_format": self.amount_format,
                        "alignment": self.align_right,
                    },
                ]
                self.excel_data.extend(invoice_line)
                j += 1
            """ write job summary line """
            sum_line = [
                {
                    "cell": f"B{9+j}",
                    "data": titles["total"],
                    "font": self.normal_font,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"C{9+j}",
                    "data": verified_amount_sum,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"D{9+j}",
                    "data": safety_deposit_sum,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"E{9+j}",
                    "data": approved_amount_sum,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
            ]
            self.excel_data.extend(sum_line)
            # add two lines, one is a spacer to the next job
            j += 2
        self.last_row_index += j

    @debug.log
    def make_file(self):
        """Make the file."""
        self.make_cells(self.excel_data)
        self.ws.print_area = f"A1:E{self.last_row_index}"
        self.save_file()


class TradesOVExcelTemplate(Template):
    """Template for an overview of the trades.

    Attributes:
        excel_data (dict): Data to write
        filename (str): Filename for output
        last_row_index (int): Index of the last used row
        save_dir (Path): Path to save directory
        save_path (Path): Path to save file
        template_dir (TYPE): Description
        template_filename (str): Name of the template-file
    """

    def __init__(self, app_data, save_dir, filename=None):
        """Initialize Template.

        Args:
            app_data (api.AppData): Application data containing the project data
            save_dir (Path): Path to save directory
            filename (None, optional): Filename for output
        """
        self.template_filename = "trades_ov.xlsx"

        self.filename = filename if filename else f"{helper.today_str()}-trades_ov"
        self.save_dir = save_dir
        self.save_path = os.path.join(
            app_data.get_dir(), self.save_dir, f"{self.filename}.xlsx"
        )

        self.last_row_index = 9  # last row used to define the printed area

        super(TradesOVExcelTemplate, self).__init__(
            template_dir=app_data.config["template_subdir"],
            template_filename=self.template_filename,
            save_path=self.save_path,
        )
        """
        #
        #   EXCEL DATA
        #   Gather the data for the invoice check and connect
        #   it to the cells.
        #
        """
        date = helper.today_str("%d.%m.%Y")
        self.excel_data = [
            {"cell": "B3", "data": date, "number_format": self.date_format},
            {"cell": "B4", "data": app_data.project.identifier},
            {
                "cell": "B5",
                "data": app_data.project.client if app_data.project.client else "",
            },
        ]

        """
        #   new_rows
        #   We need one row for each job and each invoice
        #
        """
        titles = app_data.get_titles()
        j = 0
        header_height = 30
        trades = [
            trade
            for trade in app_data.project.trades
            if len(app_data.project.get_jobs_of_trade(trade)) > 0
        ]
        for trade in trades:
            trade_lines = [
                {
                    "cell": f"A{9+j}",
                    "data": f"{trade.name}",
                    "border": Border(bottom=Side(style="medium")),
                    "font": self.bold_font_header,
                    "row_height": header_height,
                },
                {"cell": f"B{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"C{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"D{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"E{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"F{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"G{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"H{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"I{9+j}", "border": Border(bottom=Side(style="medium"))},
                {
                    "cell": f"H{10+j}",
                    "data": titles["budget"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"I{10+j}",
                    "data": titles["budget_w_VAT"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"B{10+j}",
                    "data": titles["comment"],
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"H{11+j}",
                    "data": trade.budget,
                    "font": self.normal_font,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"I{11+j}",
                    "data": trade.budget * (1 + app_data.project.get_vat()),
                    "font": self.normal_font,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                },
                {"cell": f"B{11+j}", "data": trade.comment, "font": self.normal_font},
            ]
            self.excel_data.extend(trade_lines)
            jobs_header_lines = [
                {
                    "cell": f"A{12+j}",
                    "data": titles["jobs"],
                    "border": Border(bottom=Side(style="thin", color="999999")),
                    "font": self.normal_font,
                    "row_height": header_height,
                },
                {
                    "cell": f"B{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"C{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"D{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"E{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"F{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"G{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"H{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"I{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"B{13+j}",
                    "data": titles["company"],
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"C{13+j}",
                    "data": titles["job"],
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"D{13+j}",
                    "data": titles["cost_group"],
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"E{13+j}",
                    "data": titles["job_sum_w_VAT"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"F{13+j}",
                    "data": titles["approved_invoices_w_VAT"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"G{13+j}",
                    "data": titles["safety_deposit"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"H{13+j}",
                    "data": titles["paid_safety_deposits"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                # TODO: Change title of below
                {
                    "cell": f"I{13+j}",
                    "data": titles["approved_invoices_and_psds_by_budget"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
            ]
            self.excel_data.extend(jobs_header_lines)
            j += 5

            jobs = app_data.project.get_jobs_of_trade(trade)
            job_sum_w_VAT_sum = 0
            approved_inv_sum = 0
            safety_deposit_sum = 0
            paid_safety_deposit_sum = 0
            for job in jobs:
                job_sum_w_VAT = job.job_sum_w_additions * (
                    1 + app_data.project.get_vat()
                )
                invoices_of_job = app_data.project.get_invoices_of_job(job)
                approved_inv_amount = sum(
                    invoice.approved_amount_a_discount_amount
                    for invoice in invoices_of_job
                )
                safety_deposit = sum(
                    invoice.safety_deposit_amount for invoice in invoices_of_job
                )
                paid_safety_deposits = job.paid_safety_deposits_sum
                """ sums """
                job_sum_w_VAT_sum += job_sum_w_VAT
                approved_inv_sum += approved_inv_amount
                safety_deposit_sum += safety_deposit
                paid_safety_deposit_sum += paid_safety_deposits
                """ write data """
                job_line = [
                    {"cell": f"B{9+j}", "data": job.company.name},
                    {"cell": f"C{9+j}", "data": job.id, "alignment": self.align_left},
                    {
                        "cell": f"D{9+j}",
                        "data": job.cost_group.id,
                        "alignment": self.align_left,
                    },
                    {
                        "cell": f"E{9+j}",
                        "data": job_sum_w_VAT,
                        "number_format": self.amount_format,
                        "alignment": self.align_right,
                    },
                    {
                        "cell": f"F{9+j}",
                        "data": approved_inv_amount,
                        "number_format": self.amount_format,
                        "alignment": self.align_right,
                    },
                    {
                        "cell": f"G{9+j}",
                        "data": safety_deposit,
                        "number_format": self.amount_format,
                        "alignment": self.align_right,
                    },
                    {
                        "cell": f"H{9+j}",
                        "data": paid_safety_deposits,
                        "number_format": self.amount_format,
                        "alignment": self.align_right,
                    },
                    {
                        "cell": f"I{9+j}",
                        "data": (approved_inv_amount + paid_safety_deposits)
                        / job_sum_w_VAT
                        if job_sum_w_VAT > 0
                        else 0,
                        "number_format": self.percent_format,
                        "alignment": self.align_right,
                    },
                ]
                self.excel_data.extend(job_line)
                j += 1
            """ write trade summary line """
            sum_line = [
                {
                    "cell": f"B{9+j}",
                    "data": titles["total"],
                    "font": self.normal_font,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"C{9+j}",
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"D{9+j}",
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"E{9+j}",
                    "data": job_sum_w_VAT_sum,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"F{9+j}",
                    "data": approved_inv_sum,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"G{9+j}",
                    "data": safety_deposit_sum,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"H{9+j}",
                    "data": paid_safety_deposit_sum,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"I{9+j}",
                    "data": (approved_inv_sum + paid_safety_deposit_sum)
                    / job_sum_w_VAT_sum
                    if job_sum_w_VAT_sum > 0
                    else 0,
                    "number_format": self.percent_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
            ]
            self.excel_data.extend(sum_line)
            j += 2

        self.last_row_index += j

    @debug.log
    def make_file(self):
        """Make the file."""
        self.make_cells(self.excel_data)
        self.ws.print_area = f"A1:I{self.last_row_index}"
        self.save_file()


class CostGroupsOVExcelTemplate(Template):
    """Template for an overview of the cost groups.

    Attributes:
        excel_data (dict): Data to write
        filename (str): Filename for output
        last_row_index (int): Index of the last used row
        save_dir (Path): Path to save directory
        save_path (Path): Path to save file
        template_filename (str): Name of the template-file
    """

    def __init__(self, app_data, save_dir, filename=None):
        """Initialize Template.

        Args:
            app_data (api.AppData): Application data containing the project data
            save_dir (Path): Path to save directory
            filename (None, optional): Filename for output
        """
        self.template_filename = "cost_groups_ov.xlsx"

        self.filename = filename if filename else f"{helper.today_str()}-cost_groups_ov"
        self.save_dir = save_dir
        self.save_path = os.path.join(
            app_data.get_dir(), self.save_dir, f"{self.filename}.xlsx"
        )

        self.last_row_index = 9  # last row used to define the printed area

        super(CostGroupsOVExcelTemplate, self).__init__(
            template_dir=app_data.config["template_subdir"],
            template_filename=self.template_filename,
            save_path=self.save_path,
        )
        """
        #
        #   EXCEL DATA
        #   Gather the data for the invoice check and connect
        #   it to the cells.
        #
        """
        date = helper.today_str("%d.%m.%Y")
        self.excel_data = [
            {"cell": "B3", "data": date, "number_format": self.date_format},
            {"cell": "B4", "data": app_data.project.identifier},
            {
                "cell": "B5",
                "data": app_data.project.client if app_data.project.client else "",
            },
        ]

        """
        #   new_rows
        #   We need one row for each job and each invoice
        #
        """
        titles = app_data.get_titles()
        j = 0
        header_height = 30
        cost_groups = [
            cost_group
            for cost_group in app_data.project.cost_groups
            if len(app_data.project.get_jobs_of_cost_group(cost_group)) > 0
        ]
        for cost_group in cost_groups:
            cost_group_lines = [
                {
                    "cell": f"A{9+j}",
                    "data": f"{cost_group.id} {cost_group.name}",
                    "border": Border(bottom=Side(style="medium")),
                    "font": self.bold_font_header,
                    "row_height": header_height,
                },
                {"cell": f"B{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"C{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"D{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"E{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"F{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"G{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"H{9+j}", "border": Border(bottom=Side(style="medium"))},
                {"cell": f"I{9+j}", "border": Border(bottom=Side(style="medium"))},
                {
                    "cell": f"H{10+j}",
                    "data": titles["budget"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"I{10+j}",
                    "data": titles["budget_w_VAT"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"B{10+j}",
                    "data": titles["description"],
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"H{11+j}",
                    "data": cost_group.budget,
                    "font": self.normal_font,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"I{11+j}",
                    "data": cost_group.budget * (1 + app_data.project.get_vat()),
                    "font": self.normal_font,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"B{11+j}",
                    "data": cost_group.description,
                    "font": self.normal_font,
                },
            ]
            self.excel_data.extend(cost_group_lines)
            jobs_header_lines = [
                {
                    "cell": f"A{12+j}",
                    "data": titles["jobs"],
                    "border": Border(bottom=Side(style="thin", color="999999")),
                    "font": self.normal_font,
                    "row_height": header_height,
                },
                {
                    "cell": f"B{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"C{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"D{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"E{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"F{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"G{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"H{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"I{12+j}",
                    "border": Border(bottom=Side(style="thin", color="999999")),
                },
                {
                    "cell": f"B{13+j}",
                    "data": titles["company"],
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"C{13+j}",
                    "data": titles["job"],
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"D{13+j}",
                    "data": titles["trade"],
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"E{13+j}",
                    "data": titles["job_sum_w_VAT"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"F{13+j}",
                    "data": titles["approved_invoices_w_VAT"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"G{13+j}",
                    "data": titles["safety_deposit"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                {
                    "cell": f"H{13+j}",
                    "data": titles["paid_safety_deposits"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
                # TODO: Change title of below
                {
                    "cell": f"I{13+j}",
                    "data": titles["approved_invoices_and_psds_by_budget"],
                    "font": self.bold_font_table_header,
                    "alignment": self.align_right,
                },
            ]
            self.excel_data.extend(jobs_header_lines)
            j += 5

            jobs = app_data.project.get_jobs_of_cost_group(cost_group)
            job_sum_w_VAT_sum = 0
            approved_inv_sum = 0
            safety_deposit_sum = 0
            paid_safety_deposit_sum = 0
            for job in jobs:
                job_sum_w_VAT = job.job_sum_w_additions * (
                    1 + app_data.project.get_vat()
                )
                invoices_of_job = app_data.project.get_invoices_of_job(job)
                approved_inv_amount = sum(
                    invoice.approved_amount_a_discount_amount
                    for invoice in invoices_of_job
                )
                safety_deposit = sum(
                    invoice.safety_deposit_amount for invoice in invoices_of_job
                )
                paid_safety_deposits = job.paid_safety_deposits_sum
                """ sums """
                job_sum_w_VAT_sum += job_sum_w_VAT
                approved_inv_sum += approved_inv_amount
                safety_deposit_sum += safety_deposit
                paid_safety_deposit_sum += paid_safety_deposits
                """ write data """
                job_line = [
                    {"cell": f"B{9+j}", "data": job.company.name},
                    {"cell": f"C{9+j}", "data": job.id, "alignment": self.align_left},
                    {
                        "cell": f"D{9+j}",
                        "data": job.trade.name,
                        "alignment": self.align_left,
                    },
                    {
                        "cell": f"E{9+j}",
                        "data": job_sum_w_VAT,
                        "number_format": self.amount_format,
                        "alignment": self.align_right,
                    },
                    {
                        "cell": f"F{9+j}",
                        "data": approved_inv_amount,
                        "number_format": self.amount_format,
                        "alignment": self.align_right,
                    },
                    {
                        "cell": f"G{9+j}",
                        "data": safety_deposit,
                        "number_format": self.amount_format,
                        "alignment": self.align_right,
                    },
                    {
                        "cell": f"H{9+j}",
                        "data": paid_safety_deposits,
                        "number_format": self.amount_format,
                        "alignment": self.align_right,
                    },
                    {
                        "cell": f"I{9+j}",
                        "data": (approved_inv_amount + paid_safety_deposits)
                        / job_sum_w_VAT
                        if job_sum_w_VAT > 0
                        else 0,
                        "number_format": self.percent_format,
                        "alignment": self.align_right,
                    },
                ]
                self.excel_data.extend(job_line)
                j += 1
            """ write cost_group summary line """
            sum_line = [
                {
                    "cell": f"B{9+j}",
                    "data": titles["total"],
                    "font": self.normal_font,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"C{9+j}",
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"D{9+j}",
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"E{9+j}",
                    "data": job_sum_w_VAT_sum,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"F{9+j}",
                    "data": approved_inv_sum,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"G{9+j}",
                    "data": safety_deposit_sum,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"H{9+j}",
                    "data": paid_safety_deposit_sum,
                    "number_format": self.amount_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
                {
                    "cell": f"I{9+j}",
                    "data": (approved_inv_sum + paid_safety_deposit_sum)
                    / job_sum_w_VAT_sum
                    if job_sum_w_VAT_sum > 0
                    else 0,
                    "number_format": self.percent_format,
                    "alignment": self.align_right,
                    "font": self.bold_font_table_header,
                    "border": Border(top=Side(style="double", color="999999")),
                },
            ]
            self.excel_data.extend(sum_line)
            j += 2

        self.last_row_index += j

    @debug.log
    def make_file(self):
        """Make the file."""
        self.make_cells(self.excel_data)
        self.ws.print_area = f"A1:I{self.last_row_index}"
        self.save_file()


class PCCCostGroupsOVExcelTemplate(Template):
    """Template for an overview of a project cost calculation by cost groups.

    Attributes:
        excel_data (dict): Data to write
        filename (str): Filename for output
        last_row_index (TYPE): Description
        new_rows (int): Description
        save_dir (Path): Path to save directory
        save_path (Path): Path to save file
        template_filename (str): Name of the template-file
    """

    def __init__(self, app_data, pcc, cost_groups, save_dir, filename=None):
        """Initialize Template.

        Args:
            app_data (api.AppData): Application data containing the project data
            pcc (proj.ProjectCostCalculation): The project cost calculation
            cost_groups (list): List of CostGroups to be shown
            save_dir (Path): Path to save directory
            filename (None, optional): Filename for output
        """
        self.template_filename = "pcc_cost_groups_ov.xlsx"

        self.filename = (
            filename
            if filename
            else f"{helper.today_str()}-pcc_cost_groups_ov-{app_data.project.identifier}-{pcc.name}"
        )
        self.save_dir = save_dir
        self.save_path = os.path.join(
            app_data.get_dir(), self.save_dir, f"{self.filename}.xlsx"
        )

        self.last_row_index = 31 + len(
            cost_groups
        )  # last row used to define the printed area

        super(PCCCostGroupsOVExcelTemplate, self).__init__(
            template_dir=app_data.config["template_subdir"],
            template_filename=self.template_filename,
            save_path=self.save_path,
        )

        """
        #   new_rows
        #   For each cost_group (except the first) add a new line
        #
        """
        self.new_rows = 0
        if len(cost_groups) > 1:
            self.new_rows = len(cost_groups) - 1
            self.add_rows(12, self.new_rows)

        """
        #
        #   EXCEL DATA
        #   Gather the data for the invoice check and connect
        #   it to the cells.
        #
        """
        cg_300 = [cg for cg in app_data.project.cost_groups if cg.id == "300"][0]
        cg_400 = [cg for cg in app_data.project.cost_groups if cg.id == "400"][0]
        cg_500 = [cg for cg in app_data.project.cost_groups if cg.id == "500"][0]

        projected_costs_300_400_500 = (
            pcc.get_sub_cost_groups_prognosis(cg_300, app_data.project.cost_groups)
            + pcc.get_sub_cost_groups_prognosis(cg_400, app_data.project.cost_groups)
            + pcc.get_sub_cost_groups_prognosis(cg_500, app_data.project.cost_groups)
        )
        bgf = app_data.project.project_data.usable_floor_space_bgf
        nuf = app_data.project.project_data.usable_floor_space_nuf
        self.excel_data = [
            {"cell": "A4", "data": pcc.type.split()[0]},
            {"cell": "A5", "data": pcc.name},
            {"cell": "B6", "data": app_data.project.identifier},
            {"cell": "B7", "data": helper.qdate_to_str(pcc.date)},
            {
                "cell": f"C{15+self.new_rows}",
                "data": app_data.project.get_vat(),
                "number_format": self.percent_format,
            },
            {
                "cell": f"H{14+self.new_rows}",
                "data": pcc.total_cost,
                "number_format": self.amount_format,
            },
            {
                "cell": f"H{15+self.new_rows}",
                "data": pcc.total_cost * app_data.project.get_vat(),
                "number_format": self.amount_format,
            },
            {
                "cell": f"H{16+self.new_rows}",
                "data": pcc.total_cost * (1 + app_data.project.get_vat()),
                "number_format": self.amount_format,
            },
            {"cell": f"C{19+self.new_rows}", "data": bgf},
            # TODO: was ist NRF?
            {"cell": f"C{20+self.new_rows}", "data": "TODO: was ist NRF?"},
            {"cell": f"C{21+self.new_rows}", "data": nuf},
            {
                "cell": f"D{24+self.new_rows}",
                "data": projected_costs_300_400_500 / bgf if bgf != 0 else 0,
                "number_format": self.number_format,
            },
            {
                "cell": f"G{24+self.new_rows}",
                "data": projected_costs_300_400_500 / bgf if bgf != 0 else 0,
                "number_format": self.number_format,
            },
            {
                "cell": f"D{25+self.new_rows}",
                "data": projected_costs_300_400_500 / nuf if nuf != 0 else 0,
                "number_format": self.number_format,
            },
            {
                "cell": f"G{25+self.new_rows}",
                "data": projected_costs_300_400_500 / nuf if nuf != 0 else 0,
                "number_format": self.number_format,
            },
        ]

        i = 0
        for cost_group in cost_groups:
            font = (
                self.normal_font
                if cost_group.is_sub_group()
                else self.bold_font_table_header
            )
            cost_group_prognosis = (
                pcc.get_cost_group_prognosis(cost_group)
                if cost_group.is_sub_group()
                else pcc.get_sub_cost_groups_prognosis(
                    cost_group, app_data.project.cost_groups
                )
            )
            cost_group_data = [
                {
                    "cell": f"A{12+i}",
                    "data": f"{cost_group.id} {cost_group.name}",
                    "border": Border(left=Side(style="medium"), top=Side(style="thin")),
                    "font": font,
                },
                {
                    "cell": f"G{12+i}",
                    "data": cost_group_prognosis,
                    "number_format": self.amount_format,
                    "border": Border(left=Side(style="thin"), right=Side(style="thin")),
                    "font": font,
                },
                {"cell": f"B{12+i}", "border": Border(top=Side(style="thin"))},
                {"cell": f"C{12+i}", "border": Border(top=Side(style="thin"))},
                {"cell": f"D{12+i}", "border": Border(top=Side(style="thin"))},
                {"cell": f"E{12+i}", "border": Border(top=Side(style="thin"))},
                {"cell": f"F{12+i}", "border": Border(top=Side(style="thin"))},
                {"cell": f"G{12+i}", "border": Border(top=Side(style="thin"))},
                {
                    "cell": f"H{12+i}",
                    "border": Border(
                        left=Side(style="thin"), right=Side(style="medium")
                    ),
                },
            ]
            self.excel_data.extend(cost_group_data)
            i += 1

    @debug.log
    def make_file(self):
        """Make the file."""
        self.make_cells(self.excel_data)
        self.ws.print_area = f"A1:I{self.last_row_index}"
        self.save_file()


# TODO: finish this
class PCCTradesOVExcelTemplate(Template):
    """Template for an overview of a project cost calculation by trades.

    Attributes:
        excel_data (dict): Data to write
        filename (str): Filename for output
        last_row_index (TYPE): Description
        save_dir (Path): Path to save directory
        save_path (Path): Path to save file
        template_filename (str): Name of the template-file
    """

    def __init__(self, app_data, pcc, save_dir, filename=None):
        """Initialize Template.

        Args:
            app_data (api.AppData): Application data containing the project data
            pcc (proj.ProjectCostCalculation): The project cost calculation
            save_dir (Path): Path to save directory
            filename (None, optional): Filename for output
        """
        self.template_filename = "pcc_trades_ov.xlsx"

        self.filename = (
            filename
            if filename
            else f"{helper.today_str()}-pcc_trades_ov-{app_data.project.identifier}-{pcc.name}"
        )
        self.save_dir = save_dir
        self.save_path = os.path.join(
            app_data.get_dir(), self.save_dir, f"{self.filename}.xlsx"
        )

        trades = [
            trade
            for trade in app_data.project.trades
            if pcc.get_trade_prognosis(trade) > 0
        ]
        cost_groups = [
            cost_group
            for cost_group in app_data.project.cost_groups
            if pcc.get_cost_group_prognosis(cost_group) > 0
        ]
        cost_groups = sorted(
            list(cost_groups), key=lambda cost_group: cost_group.id, reverse=False
        )

        super(PCCTradesOVExcelTemplate, self).__init__(
            template_dir=app_data.config["template_subdir"],
            template_filename=self.template_filename,
            save_path=self.save_path,
        )
        """
        #
        #   EXCEL DATA
        #   Gather the data for the invoice check and connect
        #   it to the cells.
        #
        """
        lightcolor = "EEEEEE"
        self.excel_data = [
            {"cell": "A4", "data": pcc.type.split()[0]},
            {"cell": "A5", "data": pcc.name},
            {"cell": "B6", "data": app_data.project.identifier},
            {"cell": "B7", "data": helper.qdate_to_str(pcc.date)},
        ]

        i = 0
        for cost_group in cost_groups:
            cost_group_data = [
                {
                    "cell": f"A{12+i}",
                    "data": f"{cost_group.id} {cost_group.name}",
                    "border": Border(
                        left=Side(style="medium"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    ),
                    "font": self.bold_font_table_header,
                },
                {
                    "cell": f"G{12+i}",
                    "data": pcc.get_cost_group_prognosis(cost_group),
                    "number_format": self.amount_format,
                    "border": Border(left=Side(style="thin"), right=Side(style="thin")),
                },
                {
                    "cell": f"B{12+i}",
                    "border": Border(top=Side(style="thin"), bottom=Side(style="thin")),
                },
                {
                    "cell": f"C{12+i}",
                    "border": Border(top=Side(style="thin"), bottom=Side(style="thin")),
                },
                {
                    "cell": f"D{12+i}",
                    "border": Border(top=Side(style="thin"), bottom=Side(style="thin")),
                },
                {
                    "cell": f"E{12+i}",
                    "border": Border(top=Side(style="thin"), bottom=Side(style="thin")),
                },
                {
                    "cell": f"F{12+i}",
                    "border": Border(top=Side(style="thin"), bottom=Side(style="thin")),
                },
                {
                    "cell": f"G{12+i}",
                    "border": Border(top=Side(style="thin"), bottom=Side(style="thin")),
                },
                {
                    "cell": f"H{12+i}",
                    "border": Border(
                        left=Side(style="thin"), right=Side(style="medium")
                    ),
                },
            ]
            self.excel_data.extend(cost_group_data)
            i += 1
            for trade in trades:
                if pcc.get_cost_group_trade_prognosis(cost_group, trade) > 0:
                    trade_data = [
                        {
                            "cell": f"A{12+i}",
                            "data": f"{trade.name}",
                            "border": Border(
                                left=Side(style="medium"),
                                bottom=Side(style="thin", color=lightcolor),
                            ),
                        },
                        {
                            "cell": f"B{12+i}",
                            "border": Border(
                                top=None, bottom=Side(style="thin", color=lightcolor)
                            ),
                        },
                        {
                            "cell": f"C{12+i}",
                            "border": Border(
                                top=None, bottom=Side(style="thin", color=lightcolor)
                            ),
                        },
                        {
                            "cell": f"D{12+i}",
                            "border": Border(
                                top=None, bottom=Side(style="thin", color=lightcolor)
                            ),
                        },
                        {
                            "cell": f"E{12+i}",
                            "border": Border(
                                top=None, bottom=Side(style="thin", color=lightcolor)
                            ),
                        },
                        {
                            "cell": f"F{12+i}",
                            "border": Border(
                                top=None, bottom=Side(style="thin", color=lightcolor)
                            ),
                        },
                        {
                            "cell": f"G{12+i}",
                            "data": pcc.get_cost_group_trade_prognosis(
                                cost_group, trade
                            ),
                            "number_format": self.amount_format,
                            "border": Border(
                                top=None,
                                bottom=Side(style="thin", color=lightcolor),
                                right=Side(style="thin"),
                            ),
                        },
                        {
                            "cell": f"H{12+i}",
                            "border": Border(
                                left=Side(style="thin"),
                                right=Side(style="medium"),
                                top=None,
                            ),
                        },
                    ]
                    self.excel_data.extend(trade_data)
                    i += 1

        sum_line = [
            {
                "cell": f"C{15+i-1}",
                "data": app_data.project.get_vat(),
                "number_format": self.percent_format,
            },
            {
                "cell": f"H{14+i-1}",
                "data": pcc.total_cost,
                "number_format": self.amount_format,
            },
            {
                "cell": f"H{15+i-1}",
                "data": pcc.total_cost * app_data.project.get_vat(),
                "number_format": self.amount_format,
            },
            {
                "cell": f"H{16+i-1}",
                "data": pcc.total_cost * (1 + app_data.project.get_vat()),
                "number_format": self.amount_format,
            },
        ]
        self.excel_data.extend(sum_line)

        """
        #   new_rows
        #   For each cost_group (except the first) add a new line
        #
        """
        self.add_rows(12, i - 1)
        self.last_row_index = 31 + i  # last row used to define the printed area

    @debug.log
    def make_file(self):
        """Make the file."""
        self.make_cells(self.excel_data)
        self.ws.print_area = f"A1:H{self.last_row_index}"
        self.save_file()
